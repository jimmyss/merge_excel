import sys
import queue
import ttkbootstrap as ttk
import tkinter.messagebox as messagebox
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from ttkbootstrap.constants import *
from threading import Thread
from faces import init_face
from faces.methods import re_Text


class data_validation():

    def __init__(self, master):

        """

        :param master: 主界面窗口
        """

        self.master=master
        self.master.config()
        # ---------------变量----------------
        self.select_base_sheet=ttk.StringVar()
        self.select_data_sourve=ttk.StringVar()
        self.base_select_data=['张三','李四']
        self.file_path=""
        self.work_sheet=""
        self.msg_queue = queue.Queue()

        # --------------合并excel表格界面--------------------
        self.dataVali_face = ttk.Frame(self.master,)
        self.dataVali_face.grid()

        # --------------初始化合并excel表格界面组件--------------
        # 按钮组件
        self.back = ttk.Button(self.dataVali_face, text="     返回     ",command=self.to_iniface)  # ----------------返回主界面按钮
        self.back.grid(row=0, column=2, sticky=W, padx=10, pady=10)
        self.choose_file=ttk.Button(self.dataVali_face, text="选择文件", command=self.choose_file)#-------------------选择文件按钮
        self.choose_file.grid(row=0, column=0, sticky=E, padx=10, pady=10)
        self.submit_button=ttk.Button(self.dataVali_face, text="确认并创建", command=self.submit_valid)#---------------创建数据有效性
        self.submit_button.grid(row=4, column=3, sticky=W, pady=10, padx=10)
        self.submit_button.grid_forget()

        #下拉框组件
        self.base_box=ttk.Combobox(master=self.dataVali_face, textvariable=self.select_base_sheet, font=('微软雅黑', 10),#---------------------选择工作簿
                                   values=self.base_select_data, height=15, width=30, state='normal', cursor='plus',
                                   postcommand=lambda :self.set_value_before_choose(self.base_select_data, self.select_base_sheet, self.base_box))
        self.base_box.grid(row=1, column=1, sticky=W, pady=10, padx=10)
        self.base_box.bind('<<ComboboxSelected>>', self.submit_base)
        self.base_box.grid_forget()
        self.data_source_box=ttk.Combobox(master=self.dataVali_face, textvariable=self.select_data_sourve, font=('微软雅黑', 10),
                                   values=self.base_select_data, height=15, width=30, state='normal', cursor='plus',
                                   postcommand=lambda :self.set_value_before_choose(self.base_select_data, self.select_data_sourve, self.data_source_box))
        self.data_source_box.grid(row=3, column=1, sticky=W, pady=10, padx=10)
        self.data_source_box.bind('<<ComboboxSelected>>', self.submit_ds)
        self.data_source_box.grid_forget()

        #标签组件
        self.base_box_label=ttk.Label(self.dataVali_face, text="选择工作簿：")#----------------------------选择工作簿标签
        self.base_box_label.grid(row=1, column=0, sticky=E, padx=10, pady=10)
        self.base_box_label.grid_forget()
        self.base_col_label = ttk.Label(self.dataVali_face, text="选择要操作的列：")#--------------------------选择操作列标签
        self.base_col_label.grid(row=2, column=0, sticky=E, padx=10, pady=10)
        self.base_col_label.grid_forget()
        self.data_source_label = ttk.Label(self.dataVali_face, text="选择数据源簿：")  # --------------------------选择数据源簿标签
        self.data_source_label.grid(row=3, column=0, sticky=E, padx=10, pady=10)
        self.data_source_label.grid_forget()
        self.data_source_col_label = ttk.Label(self.dataVali_face, text="选择数据源列：")  # --------------------------选择数据源列标签
        self.data_source_col_label.grid(row=4, column=0, sticky=E, padx=10, pady=10)
        self.data_source_col_label.grid_forget()
        self.operate_col_label=ttk.Label(self.dataVali_face, text="(如：A。代表选择A列)")
        self.operate_col_label.grid(row=2, column=2, sticky=W, padx=10, pady=10)
        self.operate_col_label.grid_forget()
        self.data_col_label=ttk.Label(self.dataVali_face, text="(如：A。代表选择A列)")
        self.data_col_label.grid(row=4, column=2, sticky=W, padx=10, pady=10)
        self.data_col_label.grid_forget()

        # 文本框组件
        self.file_name=ttk.Text(self.dataVali_face, height=1, width=35)#---------------------------------显示选择的文件名
        self.file_name.grid(row=0, column=1, sticky=W, padx=10, pady=10)
        self.operate_col=ttk.Text(self.dataVali_face, height=1, width=35)#-------------------------------选择操作列文本框
        self.operate_col.grid(row=2, column=1, sticky=W, pady=10, padx=10)
        self.operate_col.grid_forget()
        self.data_col=ttk.Text(self.dataVali_face, height=1, width=35)#----------------------------------选择数据源列文本框
        self.data_col.grid(row=4, column=1, sticky=W, pady=10, padx=10)
        self.data_col.grid_forget()
        self.message_box = ttk.Text(self.dataVali_face, height=10, width=35)  # ----------------------------消息框组件
        self.message_box.config(state=DISABLED)
        self.message_box.grid(row=5, column=1, sticky=W, padx=10)

        # 启动after方法
        self.master.after(100, self.show_msg)

        # 映射stdout到re_Text
        sys.stdout = re_Text(self.msg_queue)

    # --------------控制方法-------------------
    def to_iniface(self):
        self.dataVali_face.destroy()
        init_face.init_face(self.master)

    def set_message_box(self, text):
        self.message_box.config(state=NORMAL)
        self.message_box.delete('1.0', 'end')
        self.message_box.insert(END, text)
        self.message_box.config(state=DISABLED)

    def add_validation(self, operate_col, source_col):
        work_book=load_workbook(self.file_path)
        work_sheet=work_book[self.base_box.get()]#打开工作簿
        data_sheet=work_book[self.data_source_box.get()]
        print('成功打开工作簿')
        #循环把指定列设置数据有效性
        ws_max_row=work_sheet.max_row
        ds_max_row=data_sheet.max_row
        print('正在修改数据有效性')
        for i in range(2, ws_max_row+1):
            cell=operate_col.replace('\n','')+str(i)
            formuala='=OFFSET({0}!${1}$1, MATCH("*"&${2}&"*", {0}!${1}$2:${1}${3},0),, COUNTIF({0}!${1}$2:${1}${3},"*"&${2}&"*"),)'.format(self.data_source_box.get(), source_col.replace('\n',''), cell, ds_max_row)
            dv=DataValidation(type='list', formula1=formuala, allow_blank=True)
            dv.showErrorMessage = False
            dv.add(cell)
            work_sheet.add_data_validation(dv)
        work_book.save(self.file_path)
        print("修改成功，保存在："+self.file_path)
        messagebox.showinfo(title='数据有效性任务通知', message='数据有效性修改完成，保存在：'+self.file_path)


    def submit_valid(self):
        result=""
        #验证数据是否都填写
        if not self.base_box.get():
            result="请选择工作簿！\n"
            self.set_message_box(result)
            return
        if not self.data_source_box.get():
            result="请选择数据源簿！\n"
            self.set_message_box(result)
            return
        if self.operate_col.get('1.0', 'end')=='\n':
            result="请选择修改哪一列的数据有效性\n"
            self.set_message_box(result)
            return
        if self.data_col.get('1.0', 'end')=='\n':
            result="请选择修数据源所在列\n"
            self.set_message_box(result)
            return

        #若有数据进一步检查数据是否正确
        #操作表格增加数据有效性
        t=Thread(target=self.add_validation, args=(self.operate_col.get('1.0','end'), self.data_col.get('1.0','end')))
        t.start()

    def set_value_before_choose(self, base_server_data, select_base_sheet, base_box):
        """
        选择前根据文本框的内容筛选符合条件的数据
        :return:
        """

        new_select_data = []
        for i in base_server_data:
            if select_base_sheet.get() in i:  # 关键字在该选项中则追加到新的list中
                new_select_data.append(i)

        base_box["value"] = new_select_data  # 重新给下拉框赋值

    def choose_file(self):
        """

        :return: 操作消息框和合并按钮
        """
        file_path = filedialog.askopenfilename(filetypes=[("待操作文件", "*.xlsx"), ("所有文件", "*.*")])
        if file_path:  # 如果选择了文件，显示工作簿下拉框，在file_name显示当前文件路径
            self.file_name.delete('1.0', 'end')
            self.file_name.insert(END, file_path.split("/")[-1])
            self.file_path=file_path
            self.base_select_data=self.get_sheet(self.file_path)
            self.base_box.grid(row=1, column=1, sticky=W, pady=10, padx=10)#显示下拉框
            self.base_box_label.grid(row=1, column=0, sticky=E, padx=10, pady=10)#显示标签
        else:  # 如果没选择文件，但是file path中有信息，不做变动，没有信息则清空message box，隐藏合并和清空按钮
            print("no")

    def get_sheet(self, file_path):
        excel=load_workbook(file_path)
        sheets= excel.sheetnames
        return sheets

    def submit_base(self, event):
        self.operate_col.delete('1.0','end')
        self.base_col_label.grid(row=2, column=0, sticky=E, padx=10, pady=10)
        self.operate_col.grid(row=2, column=1, sticky=W, padx=10, pady=10)
        self.data_source_label.grid(row=3, column=0, sticky=E, padx=10, pady=10)
        self.data_source_box.grid(row=3, column=1, sticky=W, pady=10, padx=10)
        self.operate_col_label.grid(row=2, column=2, sticky=W, padx=10, pady=10)

    def submit_ds(self, event):
        self.data_col.delete('1.0', 'end')
        self.data_col.grid(row=4, column=1, sticky=W, padx=10, pady=10)
        self.submit_button.grid(row=4, column=3, sticky=W, pady=10, padx=10)
        self.data_source_col_label.grid(row=4, column=0, sticky=E, padx=10, pady=10)
        self.data_col_label.grid(row=4, column=2, sticky=W, padx=10, pady=10)

    def show_msg(self):
        while not self.msg_queue.empty():
            content = self.msg_queue.get()
            self.message_box.config(state=NORMAL)
            self.message_box.insert(INSERT, content)
            self.message_box.see(END)
            self.message_box.config(state=DISABLED)

        # after方法再次调用show_msg
        self.master.after(100, self.show_msg)