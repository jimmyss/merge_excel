import os
import queue
import sys
import tkinter.messagebox as messagebox
import ttkbootstrap as ttk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl import load_workbook
from ttkbootstrap.constants import *
from faces import init_face
from faces.methods import get_path, re_Text
from threading import Thread


class merge_excel():

    def __init__(self, master):
        """

        :param master: 主界面窗口
        """
        self.master=master
        self.master.config()

        #--------------变量-------------
        self.file_path=[]
        self.msg_queue=queue.Queue()

        #--------------合并excel表格界面--------------------
        self.merge_face = ttk.Frame(self.master,)
        self.merge_face.grid()

        #--------------初始化合并excel表格界面组件--------------
        #按钮组件
        self.choose_folder=ttk.Button(self.merge_face, text='选择文件夹', command=self.choose_folder)#选择文件夹按钮
        self.choose_folder.grid(row=0, column=0, sticky=E, padx=10, pady=10)
        self.confirm=ttk.Button(self.merge_face, text='确认选择并合并', command=self.start_merge_thread)#-----------确认合并按钮
        self.confirm.grid(row=2, column=1, padx=10, pady=10)
        self.confirm.grid_forget()
        self.manual_choose=ttk.Button(self.merge_face, text='手动选择文件', command=self.choose_file)#手动选择文件按钮
        self.manual_choose.grid(row=0, column=1, sticky=W, pady=10, padx=10)
        self.clear=ttk.Button(self.merge_face, text='清空所有选择', command=self.clear_files)#--------清空所有选择按钮
        self.clear.grid(row=0, column=2, sticky=W, padx=10, pady=10)
        self.clear.grid_forget()
        self.back=ttk.Button(self.merge_face, text="     返回     ", command=self.to_iniface)#----------------返回主界面按钮
        self.back.grid(row=3, column=2, sticky=W, padx=10, pady=10)


        #标签组件
        self.message_label=ttk.Label(self.merge_face, text="消息框")#--------------------------------消息框标签
        self.message_label.grid(row=1, column=0, sticky=NE, pady=10, padx=10)
        self.file_label=ttk.Label(self.merge_face, text="保存的文件名")#------------------------------保存文件名标签
        self.file_label.grid(row=3, column=0, sticky=NE, pady=10, padx=10)

        #输入框组件
        self.file_name=ttk.Entry(self.merge_face, width=35)#----------------------------------------保存文件名输入框
        self.file_name.grid(row=3, column=1, sticky=W, padx=10, pady=10)

        #文本框组件
        self.message_box=ttk.Text(self.merge_face, height=10, width=70)#----------------------------消息框组件
        self.message_box.config(state=DISABLED)
        self.message_box.grid(row=1, column=1, sticky=W, padx=10)

        #启动after方法
        self.master.after(100, self.show_msg)

        #映射stdout到re_Text
        sys.stdout = re_Text(self.msg_queue)

    #--------------控制方法-------------------
    def choose_folder(self):
        """

        :return: 操作消息框和合并按钮
        """
        folder_path = filedialog.askdirectory(title=u'选择文件', initialdir=(os.path.expanduser('H:/')))#选择文件夹路径，保存到folder_path中
        if folder_path:#如果选择了文件夹，显示合并按钮并更新message box
            self.message_box.config(state=NORMAL)
            self.message_box.delete("1.0", "end")
            self.message_box.insert(END, '待合并的excel：\n')
            self.file_path=[]
            with os.scandir(folder_path) as i:
                for entry in i:
                    if entry.is_file() and entry.name[-5:] == '.xlsx':
                        self.file_path.append(entry.path.replace("\\","/"))#将所有\改为/方便读取文件
                        self.message_box.insert(END, entry.path.replace("\\", "/")+'\n')#向message box里加入文件绝对路径
            self.message_box.config(state=DISABLED)
            if self.file_path:#如果文件路径有数据，则展示按钮
                self.confirm.grid(row=2, column=1, padx=10, pady=10)
                self.clear.grid(row=0, column=2, padx=10, pady=10)
            else:#如果文件路径为空，则不展示按钮
                self.confirm.grid_forget()
                self.clear.grid_forget()
        else:#如果没选择文件夹，file path不变
            if not self.file_path:#如果如果文件路径为空，则不展示按钮，并且清空message box
                self.message_box.config(state=NORMAL)
                self.message_box.delete('1.0', 'end')
                self.confirm.grid_forget()
                self.clear.grid_forget()
                self.message_box.config(state=DISABLED)

    def choose_file(self):
        """

        :return: 操作消息框和合并按钮
        """
        files=filedialog.askopenfilenames(filetypes=[("待合并excel文件", "*.xlsx"), ("所有文件", "*.*")])
        if files:#如果选择了文件，显示合并按钮，并更新message box
            self.message_box.config(state=NORMAL)
            self.message_box.delete("1.0", "end")
            self.message_box.insert(END, '待合并的excel：\n')
            for f in files:#遍历files
                if f not in self.file_path:#如果这个文件没有选择，就把它加进file path中
                    self.file_path.append(f)
            for f in self.file_path:#将文件路径展示到message box中
                self.message_box.insert(END, f+'\n')
            self.message_box.config(state=DISABLED)
            if self.file_path:#如果file path有数据，展示下面的按钮
                self.confirm.grid(row=2, column=1, padx=10, pady=10)
                self.clear.grid(row=0, column=2, padx=10, pady=10)
            else:#否则不展示
                self.confirm.grid_forget()
                self.clear.grid_forget()
        else:#如果没选择文件，但是file path中有信息，不做变动，没有信息则清空message box，隐藏合并和清空按钮
            if not self.file_path:
                self.message_box.config(state=NORMAL)
                self.message_box.delete('1.0', 'end')
                self.confirm.grid_forget()
                self.clear.grid_forget()
                self.message_box.config(state=DISABLED)

    def start_merge_thread(self):
        """

        :return:
        """
        file_name=self.file_name.get()
        if file_name:#如果标明了文件名，判断文件名是否合法，则执行合并程序
            if file_name[-5:] == '.xlsx':
                #开始线程
                t=Thread(target=merge, args=(self.file_path, file_name))
                t.start()
            else:#若文件名不合法，做出提示
                self.file_name.insert(END, "    起名格式为：文件名.xlsx")
                self.file_name.config(bootstyle='warning')
        else:#否则在messagebox最后提示要起名
            self.file_name.insert(0, "请为整合后的文件起名！")
            self.file_name.config(bootstyle='danger')

    def clear_files(self):
        """

        :return:清空folder_path和file_path的所有内容，并清空message_box，隐藏合并按钮和清空按钮
        """
        #清空folder_path和file_path所有内容，并清空message box，隐藏合并按钮和清空按钮
        self.file_path=[]
        self.message_box.config(state=NORMAL)
        self.message_box.delete("1.0", "end")
        self.message_box.config(state=DISABLED)
        self.confirm.grid_forget()
        self.clear.grid_forget()

    def show_msg(self):
        """

        :return:用于给messagebox同步信息
        """
        while not self.msg_queue.empty():
            content = self.msg_queue.get()
            self.message_box.config(state=NORMAL)
            self.message_box.insert(INSERT, content)
            self.message_box.see(END)
            self.message_box.config(state=DISABLED)

        # after方法再次调用show_msg
        self.master.after(500, self.show_msg)

    def to_iniface(self):
        """

        :return:返回主界面
        """
        self.merge_face.destroy()
        init_face.init_face(self.master)

def merge(file_list, save_file_name):#用于合并excel的函数
    """

    :param file_list:文件列表，保存文件的绝对路径
    :param save_file_name: 保存文件的文件名，需要是xxx.xlsx格式
    :return: 整合file_list中的文件，并以save_file_name来保存到程序所在文件夹下
    """
    try:
        xl0 = file_list[0]
        data0 = []  # 复制表头数据
        wb0 = load_workbook(filename=xl0)
        ws0 = wb0.active
        for i in range(1, ws0.max_column + 1):
            data0.append(ws0.cell(row=1, column=i).value)

        data1 = []  # 复制数据
        num = len(file_list)
        for n in range(num):
            xf = file_list[n]
            wb1 = load_workbook(filename=xf)
            ws1 = wb1.active
            for i in range(2, ws1.max_row + 1):
                list = []
                for j in range(1, ws1.max_column + 1):
                    list.append(ws1.cell(row=i, column=j))
                data1.append(list)

        #汇总表头和数据,新建保存总表
        data = []
        data.append(data0)  # 添加表头
        for l in range(len(data1)):  # 添加数据
            data.append(data1[l])
        wb = Workbook()  # 新建表
        ws = wb.active
        ws.title = 'sheet'
        for n_row in range(1, len(data) + 1):  # 写入数据
            for n_col in range(1, len(data[n_row - 1]) + 1):
                if type(data[n_row-1][n_col-1])==str:
                    ws.cell(row=n_row, column=n_col, value=str(data[n_row - 1][n_col - 1]).replace('None', ''))
                elif data[n_row-1][n_col-1].data_type=='n':
                    ws.cell(row=n_row, column=n_col, value=str(data[n_row-1][n_col-1].value).replace('None', '')).data_type='n'
                else:
                    ws.cell(row=n_row, column=n_col, value=str(data[n_row - 1][n_col - 1].value).replace('None', ''))
        wb.save(filename=save_file_name)  # 保存xlsx
        messagebox.showinfo(title='合并excel任务通知', message="文件合并完成，文件保存成功！保存到：" + get_path())
        print("文件合并完成，文件保存成功！保存到：" + get_path())
    except Exception as e:
        messagebox.showerror(title='合并excel任务通知', message="文件生成失败，请重试！")
        print("文件生成失败，请重试！")