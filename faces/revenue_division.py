import csv
import sys
import queue
import openpyxl
import ttkbootstrap as ttk
import tkinter.messagebox as messagebox
from faces import init_face
from threading import Thread
from tkinter import filedialog
from ttkbootstrap.constants import *
from faces.methods import re_Text
from openpyxl import load_workbook

class revenue_division:

    def __init__(self, master):
        """

        :param master: 主界面变量传递，本函数主要是用来生成界面的所有组件
        """
        self.master = master
        self.master.config()

        #---------------变量--------------
        self.file_path=""
        self.position=[]#专门保存特定列所在位置，如热点组在第三列，保存为3，之后按顺序读取即可生成竖表 ## 在本项目中，需要依次读取：热点组，热点组收费渠道，支付方式，日期，入账（实付+余额抵扣）
        self.msg_queue = queue.Queue()

        #--------------月度分账收入界面------------
        self.division_face=ttk.Frame(self.master)
        self.division_face.grid()

        #--------------初始化月度分账收入界面组件----------
        #按钮组件
        self.choose_file=ttk.Button(self.division_face, text="选择文件", command=self.choose_file)#--------------选择文件按钮--------------
        self.choose_file.grid(row=0, column=0, pady=10, padx=10)
        self.back=ttk.Button(self.division_face, text="    返回    ", command=self.to_iniface)#--------------返回按钮-----------------
        self.back.grid(row=0, column=2, padx=10, pady=10)
        self.create_division=ttk.Button(self.division_face, text="制作收入分账", command=self.start_thread)#--------------制作收入分账按钮----------
        self.create_division.grid(row=1, column=2, pady=10, padx=10)
        self.create_division.grid_forget()

        #标签组件
        self.save_label=ttk.Label(self.division_face, text="保存文件名")#--------------保存文件名标签------------
        self.save_label.grid(row=1, column=0, pady=10, padx=10)
        self.save_label.grid_forget()
        self.notice_label=ttk.Label(self.division_face, text="(如：分账.xlsx)")#------------保存文件提示标签
        self.notice_label.grid(row=2, column=1, padx=10, pady=10)
        self.notice_label.grid_forget()

        #输入框组件
        self.save_entry=ttk.Entry(master=self.division_face, width=35)#--------------保存文件名输入框---------
        self.save_entry.grid(row=1, column=1, padx=10, pady=10)
        self.save_entry.grid_forget()

        #消息框组件
        self.choose_file_text=ttk.Text(self.division_face, height=1, width=35)#--------------选择文件消息框-----------
        self.choose_file_text.grid(row=0, column=1, padx=10, pady=10)
        self.message_box=ttk.Text(self.division_face, height=5, width=35)#--------------结果显示消息框-----------
        self.message_box.grid(row=3, column=1, pady=10, padx=10)
        self.message_box.config(state=DISABLED)

        # 启动after方法
        self.master.after(100, self.show_msg)

        # 映射stdout到re_Text
        sys.stdout = re_Text(self.msg_queue)

    #-----------控制方法-------------

    def choose_file(self):
        """

        :return: 打开文件，选择需要分账的文件
        """
        file_path=filedialog.askopenfilename(filetypes=[("需要分账的文件", "*.csv"),["需要分账的文件","*.xlsx"]])
        if file_path: # 如果选择了文件，在选择文件消息框显示选择的文件名,显示保存文件标签和输入框，以及制作收入分账按钮
            self.file_path=file_path
            self.choose_file_text.delete('1.0', 'end')
            self.choose_file_text.insert(END, file_path.split("/")[-1])
            self.save_label.grid(row=1, column=0, pady=10, padx=10)
            self.save_entry.grid(row=1, column=1, padx=10, pady=10)
            self.create_division.grid(row=1, column=2, pady=10, padx=10)
            self.notice_label.grid(row=2, column=1, padx=10, pady=10)

    def divide(self):
        """

        :return: 分账函数，主要生成一个竖表然后用竖表生成横表，文件保存在程序所在文件夹下
        """
        if self.file_path[-4:]=='.csv':
            with open(self.file_path) as csvfile:
                print('成功打开文件')
                csv_reader = csv.reader(csvfile)  # 使用csv.reader读取csvfile中的文件
                header = next(csv_reader)        # 读取第一行每一列的标题
                print('读取每列列名完成')
                #保存关键列的位置
                self.position.append(header.index('热点组')+1)
                self.position.append(header.index('热点组收费渠道')+1)
                self.position.append(header.index('支付方式')+1)
                self.position.append(header.index('分账日期')+1)
                self.position.append(header.index('实付')+1)
                self.position.append(header.index('余额抵扣')+1)
                result_rows={}
                p=self.position#暂存特殊列所在位置
                for row in csv_reader:  # 将csv 文件中的元组依次取出放入result_rows中
                    month=row[p[3] - 1].split('/', 2)[1]#提取月份
                    if len(month)==1:#补齐month，使它长度为2，如：07
                        month='0'+month
                    r=[row[p[0]-1], row[p[1]-1], row[p[2]-1], str(row[p[3]-1].split('/',2)[0])+month]#制作键
                    s=','.join(r)#制作键
                    if s not in result_rows:#如果第一次出现
                        result_rows[s]=round(float(row[p[4]-1]),2)+round(float(row[p[5]-1]),2)#根据键 对应值
                    else:#若原来已经有了，则直接在金额上添加即可
                        result_rows[s]+=round(float(row[p[4]-1]),2)+round(float(row[p[5]-1]),2)
        elif self.file_path[-5:] == '.xlsx':
            try:
                result_rows={}
                wb = load_workbook(filename=self.file_path)
                ws = wb.active
                print('成功打开文件')
                header=[]#保存表头数据
                for i in range(1, ws.max_column + 1):
                    header.append(ws.cell(row=1, column=i).value)
                print('读取每列列名完成')
                # 保存关键列的位置
                self.position.append(header.index('热点组') + 1)
                self.position.append(header.index('热点组收费渠道') + 1)
                self.position.append(header.index('支付方式') + 1)
                self.position.append(header.index('分账日期') + 1)
                self.position.append(header.index('实付') + 1)
                self.position.append(header.index('余额抵扣') + 1)
                p = self.position  # 暂存特殊列所在位置
                for row in ws:  # 将xlsx 文件中的元组依次取出放入result_rows中
                    if row[0].row == 1:
                        continue
                    month = str(row[p[3] - 1].value).split()[0].split('-')[1]  # 提取月份
                    if len(month) == 1:  # 补齐month，使它长度为2，如：07
                        month = '0' + month
                    r = [row[p[0] - 1].value, row[p[1] - 1].value, row[p[2] - 1].value, str(row[p[3] - 1].value).split()[0].split('-')[0] + month]  # 制作键
                    s = ','.join(r)  # 制作键
                    if s not in result_rows:  # 如果第一次出现
                        result_rows[s] = round(float(row[p[4] - 1].value), 2) + round(float(row[p[5] - 1].value), 2)  # 根据键 对应值
                    else:  # 若原来已经有了，则直接在金额上添加即可
                        result_rows[s] += round(float(row[p[4] - 1].value), 2) + round(float(row[p[5] - 1].value), 2)
            except Exception as e:
                print('打开文件失败：'+self.file_path)
                print(e)
                messagebox.showerror(title='分账任务通知', message='打开文件失败'+self.file_path)
        else:
            print('目前仅支持打开 .xlsx 和 .csv 格式文件！')
            messagebox.showwarning(title='分账任务通知', message='目前仅支持打开 .xlsx 和 .csv 格式文件！')
            return
        #将result_rows中的数据依次导出到xlsx中，生成竖表
        print('生成竖表中')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        #首先导出列名
        ws.cell(1, 1, '热点组')
        ws.cell(1, 2, '热点组收费渠道')
        ws.cell(1, 3, '支付方式')
        ws.cell(1, 4, '分账日期')
        ws.cell(1, 5, '入账')
        #然后依次导出元组，元组按逗号分割，最后一列写入账金额
        i=2
        for row in result_rows.keys():
            money=result_rows[row]
            ws.cell(i, 5, money)
            for j in range(1, 5):
                if j==4:
                    ws.cell(i, j, row.split(',')[j - 1]).data_type='n'
                else:
                    ws.cell(i, j, row.split(',')[j-1])
            i+=1

        #使用竖表生成横表
        print('生成横表中')
        ws2 = wb.create_sheet('横表')
        #横表列名：年度，月份，热点组名称，#年份+月份（如：#202207）
        ws2.cell(1, 1, '年度')
        ws2.cell(1, 2, '月份')
        ws2.cell(1, 3, '热点组名称')
        year_month={}
        temp=sys.maxsize
        for k in range(2, i):#循环查找年月最小的数据
            if int(ws.cell(k, 4).value) < temp:
                temp=int(ws.cell(k, 4).value)
        for k in range(4, 29):
            if str(temp)[-2:]=='13':
                temp += 88
            ws2.cell(1, k, str(temp)).data_type='n'
            year_month[temp]=k
            temp+=1
            k+=1
        ws2.freeze_panes='D1'
        i=2
        for row in result_rows.keys():
            money=result_rows[row]
            year=row.split(',')[3][:4]
            month=row.split(',')[3][-2:]
            project=row.split(',')[0]
            ws2.cell(i, 1, year).data_type='n'
            ws2.cell(i, 2, month).data_type='n'
            ws2.cell(i, 3, project)
            ws2.cell(i, year_month[int(row.split(',')[3])], money)
            i+=1
        wb.save(self.save_entry.get())#保存表格
        messagebox.showinfo(title='分账任务通知', message='表格保存完成')
        print('表格保存完成')

    def show_msg(self):
        """

        :return:用于给messagebox同步信息
        """
        while not self.msg_queue.empty():
            content = self.msg_queue.get()
            self.message_box.config(state=NORMAL)
            self.message_box.insert(INSERT, content)
            self.message_box.see(END)
            self.message_box.config(state=DISABLED)

        # after方法再次调用show_msg
        self.master.after(500, self.show_msg)

    def start_thread(self):
        """

        :return: 在点击分账按钮后，控制生成线程
        """
        if self.save_entry.get():
            t=Thread(target=self.divide, args=())
            t.start()
        else:
            print("任务生成失败，请重试")

    def to_iniface(self):
        """

        :return:转到主界面
        """
        self.division_face.destroy()
        init_face.init_face(self.master)