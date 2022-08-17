import os
import re
import queue
import sys
from datetime import datetime
import tkinter.messagebox as messagebox
import ttkbootstrap as ttk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl import load_workbook
from ttkbootstrap.constants import *
from faces import init_face
from faces.methods import get_path, re_Text
from threading import Thread


class bill_summary():

    def __init__(self, master):
        """

        :param master:主界面窗口，生成界面所有组件
        """
        self.master=master
        self.master.config()

        #--------------------变量---------------------
        self.file_path=[]
        self.msg_queue=queue.Queue()

        # --------------月度账单汇总界面--------------------
        self.merge_face = ttk.Frame(self.master, )
        self.merge_face.grid()

        # --------------初始化月度账单汇总界面组件--------------
        # 按钮组件
        self.choose_folder = ttk.Button(self.merge_face, text='选择文件夹', command=self.choose_folder)  # 选择文件夹按钮
        self.choose_folder.grid(row=0, column=0, sticky=E, padx=10, pady=10)
        self.confirm = ttk.Button(self.merge_face, text='确认选择并合并', command=self.start_summary_thread)  # -----------确认合并按钮
        self.confirm.grid(row=2, column=1, padx=10, pady=10)
        self.confirm.grid_forget()
        self.manual_choose = ttk.Button(self.merge_face, text='手动选择文件', command=self.choose_file)  # 手动选择文件按钮
        self.manual_choose.grid(row=0, column=1, sticky=W, pady=10, padx=10)
        self.clear = ttk.Button(self.merge_face, text='清空所有选择', command=self.clear_files)  # --------清空所有选择按钮
        self.clear.grid(row=0, column=2, sticky=W, padx=10, pady=10)
        self.clear.grid_forget()
        self.back = ttk.Button(self.merge_face, text="     返回     ", command=self.to_iniface)  # ----------------返回主界面按钮
        self.back.grid(row=3, column=2, sticky=W, padx=10, pady=10)

        # 标签组件
        self.message_label = ttk.Label(self.merge_face, text="消息框")  # --------------------------------消息框标签
        self.message_label.grid(row=1, column=0, sticky=NE, pady=10, padx=10)
        self.file_label = ttk.Label(self.merge_face, text="保存的文件名")  # ------------------------------保存文件名标签
        self.file_label.grid(row=3, column=0, sticky=NE, pady=10, padx=10)

        # 输入框组件
        self.file_name = ttk.Entry(self.merge_face, width=35)  # ----------------------------------------保存文件名输入框
        self.file_name.grid(row=3, column=1, sticky=W, padx=10, pady=10)

        # 文本框组件
        self.message_box = ttk.Text(self.merge_face, height=10, width=70)  # ----------------------------消息框组件
        self.message_box.config(state=DISABLED)
        self.message_box.grid(row=1, column=1, sticky=W, padx=10)

        # 启动after方法
        self.master.after(100, self.show_msg)

        # 映射stdout到re_Text
        sys.stdout = re_Text(self.msg_queue)

    # --------------控制方法-------------------
    def summary(self):
        """

        :return:批量提取file_path中的文件内容，并整合到一个excel里，保存的excel按照file_name里面的名称确定
        """
        #首先创建汇总表保存数据
        try:
            save_book=Workbook()
            save_sheet=save_book.active
            #首先填写第一行
            save_sheet['A1'] = '结算项目'
            save_sheet['B1'] = '结算月份'
            save_sheet['C1'] = '结算金额'
            save_sheet['D1'] = '合作方名称'
            save_row=2
            for file in self.file_path:  # 批量读取文件内容
                try:
                    work_book = load_workbook(file, data_only=True)
                    work_sheet = work_book.active
                    print("成功打开工作簿:"+file.rsplit('/', 1)[1])
                    # 按行读取，识别关键字：结算项目、结算月份、结算金额、合作方名称（横向）
                    month=[]
                    project=[]
                    money=[]
                    cooperation=[]
                    for row in work_sheet:
                        for cell in row:
                            #检查有无读取到指定关键字
                            if cell.value=='结算月份':
                                month.append(cell.row)
                                month.append(cell.column)
                                #从第一个非空单元格开始，向下读取信息，直到匹配到空或者非年月格式字符
                                count=0
                                for i in range(cell.row+1, work_sheet.max_row):
                                    if work_sheet[i][month[1]-1].value:#如果有数值
                                        if not self.is_chinese(str(work_sheet[i][month[1]-1].value).replace("\t","")):
                                            month.append(str(work_sheet[i][month[1]-1].value).replace("\t",""))
                                            count=1
                                        elif count==1:
                                            break
                                    else:
                                        if count==1:#循环结束
                                            break
                            elif cell.value=='结算项目':
                                project.append(cell.row)
                                project.append(cell.column)

                            elif cell.value == '结算金额':
                                money.append(cell.row)
                                money.append(cell.column)

                            elif cell.value == '合作方名称':
                                if cell.column+1 <= work_sheet.max_column:#若读取到合作方名称，且在右边有数据，则向右读取一格
                                    if work_sheet[cell.row][cell.column].value:
                                        cooperation.append(work_sheet[cell.row][cell.column].value)
                                    else:
                                        print('请在合作方名称后写具体名称！！')
                                else:#若右边没有数据，则默认向下读取，即读取多个合作方名称
                                    cooperation.append(cell.row)
                                    cooperation.append(cell.column)
                    #结算项目
                    count = 0
                    for j in range(project[0]+1, i):
                        if work_sheet[j][project[1] - 1].value:  # 如果有值
                            project.append(work_sheet[j][project[1] - 1].value)
                            count = 1
                        elif count == 1:  # 用空值代替
                            project.append('')
                    #结算金额
                    count = 0
                    for j in range(money[0]+1, i):
                        if work_sheet[j][money[1] - 1].value:  # 如果有值
                            money.append(round(float(work_sheet[j][money[1] - 1].value), 2))
                            count = 1
                        elif count == 1:  # 用空值代替
                            money.append('')
                    #合作方名称
                    if len(cooperation) > 1:#如果是按照列读取
                        count=0
                        for j in range(cooperation[0]+1, i):
                            if work_sheet[j][cooperation[1]-1].value:
                                cooperation.append(work_sheet[j][cooperation[1]-1].value)
                                count=1
                            elif count==1:
                                cooperation.append('')
                    #批量输出到整合的excel表中
                    for i in range(2, len(month)):
                        save_sheet.cell(row=save_row, column=1, value=project[i])
                        save_sheet.cell(row=save_row, column=2, value=month[i]).data_type='n'
                        save_sheet.cell(row=save_row, column=3, value=money[i])
                        if len(cooperation)==1:
                            save_sheet.cell(row=save_row, column=4, value=cooperation[0])
                        else:
                            save_sheet.cell(row=save_row, column=4, value=cooperation[i])
                        save_row+=1
                except Exception as e:
                    print("读取文件失败："+file)
                    print(e)
                    messagebox.showerror(title='账单汇总任务通知', message='读取文件失败：\n'+file)
            save_book.save(self.file_name.get())
            print('\n文件保存成功，请前往小程序所在文件夹查看！')
            messagebox.showinfo(title='账单汇总任务通知', message='账单汇总完成！请前往小程序所在文件夹查看！')
        except:
            print("文件保存失败，请重试")
            messagebox.showerror(title='账单汇总任务通知', message='文件保存失败，请重试!')

    def validate(self, date_text):
        """

        :param date_text: 需要判断合规的日期文字
        :return: 若时间格式正确，返回Ture；若不正确，返回False
        """
        try:
            if date_text != datetime.strptime(date_text, "%Y-%m").strftime('%Y-%m'):
                raise ValueError
            return True
        except ValueError:
            # raise ValueError("错误是日期格式或日期,格式是年-月-日")
            return False

    def is_chinese(self, string):
        """
        检查整个字符串是否包含中文
        :param string: 需要检查的字符串
        :return: bool
        """
        for ch in string:
            if u'\u4e00' <= ch <= u'\u9fff':
                return True

        return False

    def choose_folder(self):
        """

        :return: 操作消息框和合并按钮
        """
        folder_path = filedialog.askdirectory(title=u'选择文件',
                                              initialdir=(os.path.expanduser('H:/')))  # 选择文件夹路径，保存到folder_path中
        if folder_path:  # 如果选择了文件夹，显示合并按钮并更新message box
            self.message_box.config(state=NORMAL)
            self.message_box.delete("1.0", "end")
            self.message_box.insert(END, '待汇总的excel：\n')
            self.file_path = []
            with os.scandir(folder_path) as i:
                for entry in i:
                    if entry.is_file() and entry.name[-5:] == '.xlsx':
                        self.file_path.append(entry.path.replace("\\", "/"))
                        self.message_box.insert(END, entry.path.replace("\\", "/") + '\n')
            self.message_box.config(state=DISABLED)
            if self.file_path:
                self.confirm.grid(row=2, column=1, padx=10, pady=10)
                self.clear.grid(row=0, column=2, padx=10, pady=10)
            else:
                self.confirm.grid_forget()
                self.clear.grid_forget()
        else:  # 如果没选择文件夹，file path不变
            if not self.file_path:
                self.message_box.config(state=NORMAL)
                self.message_box.delete('1.0', 'end')
                self.confirm.grid_forget()
                self.clear.grid_forget()
                self.message_box.config(state=DISABLED)

    def choose_file(self):
        """

        :return: 操作消息框和合并按钮
        """
        files = filedialog.askopenfilenames(filetypes=[("待合并excel文件", "*.xlsx"), ("所有文件", "*.*")])
        if files:  # 如果选择了文件，显示合并按钮，并更新message box
            self.message_box.config(state=NORMAL)
            self.message_box.delete("1.0", "end")
            self.message_box.insert(END, '待合并的excel：\n')
            for f in files:
                if f not in self.file_path:
                    self.file_path.append(f)
            for f in self.file_path:
                self.message_box.insert(END, f + '\n')
            self.message_box.config(state=DISABLED)
            if self.file_path:
                self.confirm.grid(row=2, column=1, padx=10, pady=10)
                self.clear.grid(row=0, column=2, padx=10, pady=10)
            else:
                self.confirm.grid_forget()
                self.clear.grid_forget()
        else:  # 如果没选择文件，但是file path中有信息，不做变动，没有信息则清空message box，隐藏合并和清空按钮
            if not self.file_path:
                self.message_box.config(state=NORMAL)
                self.message_box.delete('1.0', 'end')
                self.confirm.grid_forget()
                self.clear.grid_forget()
                self.message_box.config(state=DISABLED)

    def clear_files(self):
        """

        :return: 清空folder_path和file_path所有内容，并清空message box，隐藏合并按钮和清空按钮
        """
        self.file_path = []
        self.message_box.config(state=NORMAL)
        self.message_box.delete("1.0", "end")
        self.message_box.config(state=DISABLED)
        self.confirm.grid_forget()
        self.clear.grid_forget()

    def show_msg(self):
        """

        :return: 用于给message_box同步信息
        """
        while not self.msg_queue.empty():
            content = self.msg_queue.get()
            self.message_box.config(state=NORMAL)
            self.message_box.insert(INSERT, content)
            self.message_box.see(END)
            self.message_box.config(state=DISABLED)

        # after方法再次调用show_msg
        self.master.after(500, self.show_msg)

    def start_summary_thread(self):
        """

        :return: 判断文件名是否填写，并生成任务线程
        """
        file_name=self.file_name.get()
        if file_name:#如果标明了文件名，判断文件名是否合法，则执行合并程序
            if file_name[-5:] == '.xlsx':
                #开始线程
                t=Thread(target=self.summary, args=())
                t.start()
            else:#若文件名不合法，做出提示
                self.file_name.insert(END, "    起名格式为：文件名.xlsx")
                self.file_name.config(bootstyle='warning')
        else:#否则在messagebox最后提示要起名
            self.file_name.insert(0, "请为整合后的文件起名！")
            self.file_name.config(bootstyle='danger')

    def to_iniface(self):
        """

        :return:返回主界面
        """
        self.merge_face.destroy()
        init_face.init_face(self.master)