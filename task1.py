# -*- coding: utf-8 -*-
import os
import sys
import copy
import xlrd
import xlsxwriter
import tkinter as tk
import ttkbootstrap as ttk
from tkinter import filedialog
from ttkbootstrap.constants import *
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def get_path():
    app_path = ""
    if getattr(sys, 'frozen', False):
        app_path = os.path.dirname(sys.executable).replace("\\", "/")
    elif __file__:
        app_path = os.path.dirname(os.path.abspath(__file__)).replace("\\", "/")
    return app_path

def merge(file_list, save_file_name):#用于合并excel的函数
    #获取保存路径
    save_path=get_path()

    if not os.path.exists(save_path + '\\' + save_file_name):
        wb = Workbook()
        # ws = wb.active
        wb.save(save_file_name)
    wbo = load_workbook(save_file_name)
    wso = wbo.active
    x=0#单元格定位x
    # 保存表头，默认以第一张表的表头为准
    title_excel = load_workbook(file_list[0])
    sheet = title_excel.active
    for i in range(1, sheet.max_column+1):
        wso.cell(1, i, sheet.cell(1, i).value)

    for f in file_list:
        wbt = load_workbook(f)
        wst = wbt.active
        max_row = wst.max_row
        max_column = wst.max_column
        for i in range(2,max_row+1):
            for j in range(1,max_column+1):
                v = wst.cell(i,j).value
                if v:
                    wso.cell(x+i,j,v)
                    if wst.cell(i,j).has_style:
                        wso.cell(x+1, j)._style = copy.copy(wst.cell(i,j)._style)
                        wso.cell(x+1, j).font = copy.copy(wst.cell(i,j).font)
                        wso.cell(x+1, j).border = copy.copy(wst.cell(i,j).border)
                        wso.cell(x+1, j).fill = copy.copy(wst.cell(i,j).fill)
                        wso.cell(x+1, j).number_format = copy.copy(wst.cell(i,j).number_format)
                        wso.cell(x+1, j).protection = copy.copy(wst.cell(i,j).protection)
                        wso.cell(x+1, j).alignment = copy.copy(wst.cell(i,j).alignment)

        x=max_row-1
    wbo.save(save_file_name)
    print('合并完成...')



class basedesk():#主界面
    def __init__(self, master):
        self.root = master
        self.root.config()
        self.root.title('Base page')
        self.root.geometry('750x400')

        initface(self.root)


class initface():
    def __init__(self, master):
        """

        :param master: 主界面窗口
        """
        self.master = master
        self.master.config()
        # 基准界面initface
        self.initface = ttk.Frame(self.master, )
        self.initface.grid(row=0, column=0)

        #初始化基准界面组件
        task1_but = ttk.Button(self.initface, text='excel合并功能', command=self.to_merge_excel)
        task1_but.grid(row=1, column=1, sticky=E, pady=10)
        task2_but = ttk.Button(self.initface, text='功能2', )
        task2_but.grid(row=2, column=1, sticky=E, pady=10)

    def to_merge_excel(self, ):
        self.initface.destroy()
        merge_excel(self.master)

class merge_excel():

    def __init__(self, master):
        """

        :param master: 主界面窗口
        """
        self.master=master
        self.master.config()

        #--------------变量-------------
        self.file_path=[]

        #--------------合并excel表格界面--------------------
        self.merge_face = ttk.Frame(self.master,)
        self.merge_face.grid()

        #--------------初始化合并excel表格界面组件--------------
        #按钮组件
        self.choose_folder=ttk.Button(self.merge_face, text='选择文件夹', command=self.choose_folder)#选择文件夹按钮
        self.choose_folder.grid(row=0, column=0, sticky=E, padx=10, pady=10)
        self.confirm=ttk.Button(self.merge_face, text='确认选择并合并', command=self.merge)#-----------确认合并按钮
        self.confirm.grid(row=2, column=1, padx=10, pady=10)
        self.confirm.grid_forget()
        self.manual_choose=ttk.Button(self.merge_face, text='手动选择文件', command=self.choose_file)#手动选择文件按钮
        self.manual_choose.grid(row=0, column=1, sticky=W, pady=10, padx=10)
        self.clear=ttk.Button(self.merge_face, text='清空所有选择', command=self.clear_files)#--------清空所有选择按钮
        self.clear.grid(row=0, column=2, sticky=W, padx=10, pady=10)
        self.clear.grid_forget()

        #标签组件
        self.message_label=ttk.Label(self.merge_face, text="消息框")#--------------------------------消息框标签
        self.message_label.grid(row=1, column=0, sticky=NE, pady=10, padx=10)
        self.file_label=ttk.Label(self.merge_face, text="保存的文件名")#------------------------------保存文件名标签
        self.file_label.grid(row=3, column=0, sticky=NE, pady=10, padx=10)
        #输入框组件
        self.file_name=ttk.Entry(self.merge_face, width=35)#----------------------------------------保存文件名输入框
        self.file_name.grid(row=3, column=1, sticky=W, padx=10, pady=10)

        #文本框组件
        self.message_box=ttk.Text(self.merge_face, height=10, width=70)#----------------------------消息框组件
        self.message_box.config(state=DISABLED)
        self.message_box.grid(row=1, column=1, sticky=W, padx=10)

        #--------------控制方法-------------------
    def choose_folder(self):
        """

        :return: 操作消息框和合并按钮
        """
        folder_path = filedialog.askdirectory(title=u'选择文件', initialdir=(os.path.expanduser('H:/')))#选择文件夹路径，保存到folder_path中
        if folder_path:#如果选择了文件夹，显示合并按钮并更新message box
            self.message_box.config(state=NORMAL)
            self.message_box.delete("1.0", "end")
            self.message_box.insert(END, '待合并的excel：\n')
            self.file_path=[]
            with os.scandir(folder_path) as i:
                for entry in i:
                    if entry.is_file() and entry.name[-5:] == '.xlsx':
                        self.file_path.append(entry.path.replace("\\","/"))
                        self.message_box.insert(END, entry.path.replace("\\", "/")+'\n')
            self.message_box.config(state=DISABLED)
            if self.file_path:
                self.confirm.grid(row=2, column=1, padx=10, pady=10)
                self.clear.grid(row=0, column=2, padx=10, pady=10)
            else:
                self.confirm.grid_forget()
                self.clear.grid_forget()
        else:#如果没选择文件夹，file path不变
            if not self.file_path:
                self.message_box.config(state=NORMAL)
                self.message_box.delete('1.0', 'end')
                self.confirm.grid_forget()
                self.clear.grid_forget()
                self.message_box.config(state=DISABLED)

    def choose_file(self):
        """

        :return: 操作消息框和合并按钮
        """
        files=filedialog.askopenfilenames(filetypes=[("待合并excel文件", "*.xlsx"), ("所有文件", "*.*")])
        if files:#如果选择了文件，显示合并按钮，并更新message box
            self.message_box.config(state=NORMAL)
            self.message_box.delete("1.0", "end")
            self.message_box.insert(END, '待合并的excel：\n')
            for f in files:
                if f not in self.file_path:
                    self.file_path.append(f)
            for f in self.file_path:
                self.message_box.insert(END, f+'\n')
            self.message_box.config(state=DISABLED)
            if self.file_path:
                self.confirm.grid(row=2, column=1, padx=10, pady=10)
                self.clear.grid(row=0, column=2, padx=10, pady=10)
            else:
                self.confirm.grid_forget()
                self.clear.grid_forget()
        else:#如果没选择文件，但是file path中有信息，不做变动，没有信息则清空message box，隐藏合并和清空按钮
            if not self.file_path:
                self.message_box.config(state=NORMAL)
                self.message_box.delete('1.0', 'end')
                self.confirm.grid_forget()
                self.clear.grid_forget()
                self.message_box.config(state=DISABLED)

    def merge(self):
        file_name=self.file_name.get()
        if file_name:#如果标明了文件名，判断文件名是否合法，则执行合并程序
            if file_name[-5:] == '.xlsx':
                merge(self.file_path, file_name)
            else:#若文件名不合法，做出提示
                self.file_name.insert(END, "    起名格式为：文件名.xlsx")
                self.file_name.config(bootstyle='warning')
        else:#否则在messagebox最后提示要起名
            self.file_name.insert(0, "请为整合后的文件起名！")
            self.file_name.config(bootstyle='danger')


    def clear_files(self):
        #清空folder_path和file_path所有内容，并清空message box，隐藏合并按钮和清空按钮
        self.file_path=[]
        self.message_box.config(state=NORMAL)
        self.message_box.delete("1.0", "end")
        self.message_box.config(state=DISABLED)
        self.confirm.grid_forget()
        self.clear.grid_forget()


if __name__ == '__main__':
    root = tk.Tk()
    basedesk(root)
    root.mainloop()